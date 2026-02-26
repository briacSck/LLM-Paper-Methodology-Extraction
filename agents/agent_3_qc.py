"""
Agent 3 — Quality Control.

Cross-validates extracted values against 11 logical/quality rules, auto-corrects
logic violations (rules 1–4), flags data-quality issues (rules 5–11), and produces
a human review queue as both JSON and Excel.
"""

from __future__ import annotations

import json
import logging
import sys
from collections import Counter
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Callable, Optional

_AGENTS_DIR = Path(__file__).resolve().parent
_PROJECT_ROOT = _AGENTS_DIR.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

import config
from schemas.extraction_schema import EXCEL_FIELDS  # noqa: F401 (confirms schema import)

import openpyxl
import openpyxl.utils
from openpyxl.styles import Font

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Flag / tag constants
# ---------------------------------------------------------------------------

FLAG_QC_MISSING_LOGIC     = "[FLAG-QC-MISSING-LOGIC]"
FLAG_QC_MEDIATOR_LOGIC    = "[FLAG-QC-MEDIATOR-LOGIC]"
FLAG_QC_MODERATOR_LOGIC   = "[FLAG-QC-MODERATOR-LOGIC]"
FLAG_QC_ENDOGENEITY_LOGIC = "[FLAG-QC-ENDOGENEITY-LOGIC]"
FLAG_QC_REPLICATION_LOGIC = "[FLAG-QC-REPLICATION-LOGIC]"
FLAG_LPM                  = "[FLAG-LPM]"
FLAG_QC_DV_VAGUE          = "[FLAG-QC-DV-VAGUE]"
FLAG_QC_IV_VAGUE          = "[FLAG-QC-IV-VAGUE]"
FLAG_QC_SAMPLE_INVALID    = "[FLAG-QC-SAMPLE-INVALID]"
FLAG_QC_CONTROL_COUNT     = "[FLAG-QC-CONTROL-COUNT]"

# Info tags — logged but do NOT trigger human review
INFO_MISSING_HANDLING_NR  = "[INFO-MISSING-HANDLING-NR]"

# ---------------------------------------------------------------------------
# Dependent-field tuples
# ---------------------------------------------------------------------------

_MISSING_DEPENDENTS = (
    "Missing_Rate_Reported", "Missing_Rate_Value", "Missing_Variables",
    "Missing_Handling", "Missing_Handling_Other", "Missing_Justified",
    "Missing_Pattern_Tested", "Missing_Pattern_Result", "Missing_Sensitivity",
)
_MEDIATOR_DEPENDENTS = (
    "Mediator_Name", "Mediator_Construct", "Mediator_Measurement", "Mediation_Method",
)
_MODERATOR_DEPENDENTS = (
    "Moderator_Name", "Moderator_Construct", "Moderator_Measurement", "Moderation_Method",
)
_ENDOGENEITY_DEPENDENTS = ("Endogeneity_Method",)

# ---------------------------------------------------------------------------
# Flag → fields mapping (for review cards)
# ---------------------------------------------------------------------------

_FLAG_TO_FIELDS: dict[str, list[str]] = {
    FLAG_QC_MISSING_LOGIC: [
        "Missing_Mentioned", "Missing_Rate_Reported", "Missing_Handling",
        "Missing_Justified", "Missing_Pattern_Tested", "Missing_Sensitivity",
    ],
    FLAG_QC_MEDIATOR_LOGIC: [
        "Mediator_Present", "Mediator_Name", "Mediator_Construct",
        "Mediator_Measurement", "Mediation_Method",
    ],
    FLAG_QC_MODERATOR_LOGIC: [
        "Moderator_Present", "Moderator_Name", "Moderator_Construct",
        "Moderator_Measurement", "Moderation_Method",
    ],
    FLAG_QC_ENDOGENEITY_LOGIC: ["Endogeneity_Addressed", "Endogeneity_Method"],
    FLAG_QC_REPLICATION_LOGIC: ["Replication_Feasibility", "Data_Available", "Code_Available"],
    FLAG_LPM:                  ["DV_Type", "Model_Type"],
    FLAG_QC_DV_VAGUE:          ["DV_Measurement"],
    FLAG_QC_IV_VAGUE:          ["IV_Measurement"],
    FLAG_QC_SAMPLE_INVALID:    ["Sample_Size"],
    FLAG_QC_CONTROL_COUNT:     ["Control_Num", "Control_List"],
}

# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------


@dataclass
class QCRule:
    name: str
    flag: str
    check: Callable[[dict], bool]
    affected_fields: list[str]
    auto_correct_fields: Optional[tuple]
    decision_question: str
    info_only: bool = False


@dataclass
class QCResult:
    paper_id: str
    qc_flags: list[str]
    info_tags: list[str]
    requires_human_review: bool
    auto_corrections: dict
    qc_passed: bool
    final_confidence: int = 2


# ---------------------------------------------------------------------------
# Private helpers — MUST be defined before _RULES
# ---------------------------------------------------------------------------


def _is_sample_invalid(d: dict) -> bool:
    ss = d.get("Sample_Size", "NR")
    if ss in ("NR", "NA", ""):
        return False
    exc_str: str = ""
    try:
        return int(ss) <= 0
    except (ValueError, TypeError) as exc:
        exc_str = str(exc)
        return True


def _check_control_count(d: dict) -> bool:
    num = d.get("Control_Num", "NR")
    lst = d.get("Control_List", "NR")
    if num in ("NR", "NA", "") or lst in ("NR", "NA", ""):
        return False
    exc_str: str = ""
    try:
        n = int(num)
        parts = [p.strip() for p in lst.split(",") if p.strip()]
        return n != len(parts)
    except (ValueError, TypeError) as exc:
        exc_str = str(exc)
        return False


# ---------------------------------------------------------------------------
# Rules list
# ---------------------------------------------------------------------------

_RULES: list[QCRule] = [
    # Rule 1
    QCRule(
        "missing_data_logic", FLAG_QC_MISSING_LOGIC,
        check=lambda d: d.get("Missing_Mentioned") == "0"
            and any(d.get(f) != "NA" for f in _MISSING_DEPENDENTS),
        affected_fields=["Missing_Mentioned"] + list(_MISSING_DEPENDENTS),
        auto_correct_fields=_MISSING_DEPENDENTS,
        decision_question=(
            "Missing_Mentioned=0 but dependent fields not 'NA'. "
            "Verify missing data was truly unmentioned."
        ),
    ),
    # Rule 2
    QCRule(
        "mediator_logic", FLAG_QC_MEDIATOR_LOGIC,
        check=lambda d: d.get("Mediator_Present") == "0"
            and any(d.get(f) != "NA" for f in _MEDIATOR_DEPENDENTS),
        affected_fields=["Mediator_Present"] + list(_MEDIATOR_DEPENDENTS),
        auto_correct_fields=_MEDIATOR_DEPENDENTS,
        decision_question=(
            "Mediator_Present=0 but mediator fields not 'NA'. Verify mediation structure."
        ),
    ),
    # Rule 3
    QCRule(
        "moderator_logic", FLAG_QC_MODERATOR_LOGIC,
        check=lambda d: d.get("Moderator_Present") == "0"
            and any(d.get(f) != "NA" for f in _MODERATOR_DEPENDENTS),
        affected_fields=["Moderator_Present"] + list(_MODERATOR_DEPENDENTS),
        auto_correct_fields=_MODERATOR_DEPENDENTS,
        decision_question=(
            "Moderator_Present=0 but moderator fields not 'NA'. Verify moderation structure."
        ),
    ),
    # Rule 4
    QCRule(
        "endogeneity_logic", FLAG_QC_ENDOGENEITY_LOGIC,
        check=lambda d: d.get("Endogeneity_Addressed") == "0"
            and d.get("Endogeneity_Method") != "NA",
        affected_fields=["Endogeneity_Addressed", "Endogeneity_Method"],
        auto_correct_fields=_ENDOGENEITY_DEPENDENTS,
        decision_question=(
            "Endogeneity_Addressed=0 but Endogeneity_Method is not 'NA'."
        ),
    ),
    # Rule 5 — flag only, no auto-correct (ambiguous which direction to fix)
    QCRule(
        "replication_logic", FLAG_QC_REPLICATION_LOGIC,
        check=lambda d: d.get("Replication_Feasibility") == "1"
            and d.get("Data_Available") not in ("1", "3"),
        affected_fields=["Replication_Feasibility", "Data_Available"],
        auto_correct_fields=None,
        decision_question=(
            "Replication_Feasibility=1 but Data_Available is not 1 or 3. "
            "Correct Replication_Feasibility or Data_Available."
        ),
    ),
    # Rule 6
    QCRule(
        "lpm_check", FLAG_LPM,
        check=lambda d: d.get("DV_Type") == "2" and d.get("Model_Type") == "1",
        affected_fields=["DV_Type", "Model_Type"],
        auto_correct_fields=None,
        decision_question=(
            "DV_Type=2 (binary) + Model_Type=1 (OLS) = linear probability model. "
            "Should Model_Type be 2 (Logit/Probit)?"
        ),
    ),
    # Rule 7
    QCRule(
        "dv_vague", FLAG_QC_DV_VAGUE,
        check=lambda d: d.get("DV_Measurement", "NR") not in ("NR", "NA", "")
            and len(str(d.get("DV_Measurement", ""))) < 25,
        affected_fields=["DV_Measurement"],
        auto_correct_fields=None,
        decision_question=(
            "DV_Measurement < 25 chars — too vague for replication. "
            "Provide formula + data source."
        ),
    ),
    # Rule 8
    QCRule(
        "iv_vague", FLAG_QC_IV_VAGUE,
        check=lambda d: d.get("IV_Measurement", "NR") not in ("NR", "NA", "")
            and len(str(d.get("IV_Measurement", ""))) < 25,
        affected_fields=["IV_Measurement"],
        auto_correct_fields=None,
        decision_question=(
            "IV_Measurement < 25 chars — too vague for replication. "
            "Provide formula + data source."
        ),
    ),
    # Rule 9
    QCRule(
        "sample_size_invalid", FLAG_QC_SAMPLE_INVALID,
        check=_is_sample_invalid,
        affected_fields=["Sample_Size"],
        auto_correct_fields=None,
        decision_question="Sample_Size is not a valid positive integer. Verify and correct.",
    ),
    # Rule 10 — INFO ONLY
    QCRule(
        "missing_nr_check", INFO_MISSING_HANDLING_NR,
        check=lambda d: d.get("Missing_Mentioned") == "1"
            and d.get("Missing_Handling") == "NR",
        affected_fields=["Missing_Mentioned", "Missing_Handling"],
        auto_correct_fields=None,
        decision_question="",
        info_only=True,
    ),
    # Rule 11
    QCRule(
        "control_count_mismatch", FLAG_QC_CONTROL_COUNT,
        check=_check_control_count,
        affected_fields=["Control_Num", "Control_List"],
        auto_correct_fields=None,
        decision_question=(
            "Control_Num does not match item count in Control_List. "
            "Verify both fields."
        ),
    ),
]

# ---------------------------------------------------------------------------
# Core QC function
# ---------------------------------------------------------------------------


def run_qc(
    extraction: dict,
    classification: Optional[dict] = None,
) -> QCResult:
    """Run all QC rules on a single extraction dict.

    Does NOT mutate the input dict. Returns a QCResult whose
    auto_corrections and final_confidence can be applied back to
    the extraction JSON by the caller.
    """
    data = dict(extraction)
    qc_flags: list[str] = []
    info_tags: list[str] = []
    auto_corrections: dict = {}

    for rule in _RULES:
        if rule.check(data):
            if rule.info_only:
                info_tags.append(rule.flag)
            else:
                qc_flags.append(rule.flag)

            if rule.auto_correct_fields is not None:
                for f in rule.auto_correct_fields:
                    if data.get(f) != "NA":
                        auto_corrections[f] = {"old": data.get(f), "new": "NA"}
                        data[f] = "NA"

    # Raise Extraction_Confidence 2 → 3 if no QC flags
    if not qc_flags:
        data["Extraction_Confidence"] = max(int(data.get("Extraction_Confidence", 2)), 3)

    final_confidence: int = int(data.get("Extraction_Confidence", 2))

    # Classifier-level review triggers
    cls_review: bool = False
    mm_trigger: bool = False
    multistudy_trigger: bool = False
    if classification is not None:
        cls_review = bool(classification.get("Requires_Human_Review", False))
        mm_trigger = classification.get("Classification_Code") == "MM"
        multistudy_trigger = "[FLAG-MULTISTUDY]" in classification.get("Flag", "")

    # Confidence trigger — papers with parse errors stay ≤ 2
    confidence_trigger: bool = final_confidence <= 2

    requires_human_review = bool(
        qc_flags or cls_review or mm_trigger or multistudy_trigger or confidence_trigger
    )
    qc_passed = not requires_human_review and not qc_flags

    return QCResult(
        paper_id=str(data.get("Paper_ID", "unknown")),
        qc_flags=qc_flags,
        info_tags=info_tags,
        requires_human_review=requires_human_review,
        auto_corrections=auto_corrections,
        qc_passed=qc_passed,
        final_confidence=final_confidence,
    )


# ---------------------------------------------------------------------------
# Batch QC
# ---------------------------------------------------------------------------


def run_qc_all(
    extractions_dir: Optional[Path] = None,
    output_dir: Optional[Path] = None,
) -> list[QCResult]:
    """Run QC on all *_extraction.json files in extractions_dir.

    For each paper:
    - Applies auto-corrections to the extraction JSON.
    - Updates Extraction_Confidence from the corrected value.
    - Embeds the QCResult as a "qc_results" key and saves back to disk.

    Returns the full list of QCResult objects.
    """
    if extractions_dir is None:
        extractions_dir = config.EXTRACTIONS_DIR
    if output_dir is None:
        output_dir = config.OUTPUT_DIR

    extractions_dir = Path(extractions_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    extraction_files = sorted(extractions_dir.glob("*_extraction.json"))
    results: list[QCResult] = []

    for ext_path in extraction_files:
        paper_id = ext_path.name.replace("_extraction.json", "")

        # Load extraction dict
        exc_str: str = ""
        try:
            extraction_data = json.loads(ext_path.read_text(encoding="utf-8"))
        except Exception as exc:
            exc_str = str(exc)
            logger.error("Failed to load %s: %s", ext_path.name, exc_str)
            continue

        # Load classification dict (optional)
        classification_data: Optional[dict] = None
        cls_path = extractions_dir / f"{paper_id}_classification.json"
        if cls_path.exists():
            exc_str = ""
            try:
                classification_data = json.loads(cls_path.read_text(encoding="utf-8"))
            except Exception as exc:
                exc_str = str(exc)
                logger.warning(
                    "Failed to load classification for %s: %s", paper_id, exc_str
                )

        # Run QC
        result = run_qc(extraction_data, classification_data)
        results.append(result)

        # Apply auto-corrections to the extraction dict
        for f, correction in result.auto_corrections.items():
            extraction_data[f] = correction["new"]

        # Sync Extraction_Confidence with post-QC value
        extraction_data["Extraction_Confidence"] = result.final_confidence

        # Embed QC results
        extraction_data["qc_results"] = asdict(result)

        # Save updated extraction JSON
        exc_str = ""
        try:
            ext_path.write_text(
                json.dumps(extraction_data, indent=2, ensure_ascii=False),
                encoding="utf-8",
            )
        except Exception as exc:
            exc_str = str(exc)
            logger.error(
                "Failed to save updated extraction for %s: %s", paper_id, exc_str
            )

    return results


# ---------------------------------------------------------------------------
# Review queue generation
# ---------------------------------------------------------------------------


def generate_review_queue(
    all_qc_results: list[QCResult],
    extractions_dir: Path,
    output_dir: Path,
) -> list[dict]:
    """Build human review artifacts (JSON + Excel) for flagged papers.

    Returns the list of review card dicts.
    """
    extractions_dir = Path(extractions_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    flagged = [r for r in all_qc_results if r.requires_human_review]

    # Build a {flag: decision_question} lookup from the rules list
    flag_to_question: dict[str, str] = {
        rule.flag: rule.decision_question for rule in _RULES
    }

    review_cards: list[dict] = []

    for result in flagged:
        paper_id = result.paper_id

        # Load extraction dict
        exc_str: str = ""
        extraction_data: dict = {}
        ext_path = extractions_dir / f"{paper_id}_extraction.json"
        if ext_path.exists():
            try:
                extraction_data = json.loads(ext_path.read_text(encoding="utf-8"))
            except Exception as exc:
                exc_str = str(exc)
                logger.warning(
                    "Could not load extraction for review card %s: %s", paper_id, exc_str
                )

        # Load classification dict
        cls_data: dict = {}
        cls_path = extractions_dir / f"{paper_id}_classification.json"
        if cls_path.exists():
            exc_str = ""
            try:
                cls_data = json.loads(cls_path.read_text(encoding="utf-8"))
            except Exception as exc:
                exc_str = str(exc)
                logger.warning(
                    "Could not load classification for review card %s: %s", paper_id, exc_str
                )

        # Build deduplicated fields_to_check list
        fields_to_check: list[str] = []
        for flag in result.qc_flags:
            for f in _FLAG_TO_FIELDS.get(flag, []):
                if f not in fields_to_check:
                    fields_to_check.append(f)

        # Build numbered decision_required string
        decision_parts: list[str] = []
        for i, flag in enumerate(result.qc_flags, 1):
            q = flag_to_question.get(flag, "")
            if q:
                decision_parts.append(f"{i}. {flag} — {q}")
        decision_required = "\n".join(decision_parts)

        # Extracted values for the relevant fields
        extracted_values = {f: extraction_data.get(f, "N/A") for f in fields_to_check}

        flags_str = (
            ", ".join(result.qc_flags)
            if result.qc_flags
            else "(no QC flags — review triggered by classifier/confidence)"
        )

        card: dict = {
            "Paper_ID": paper_id,
            "Title": extraction_data.get("Title", "N/A"),
            "Classification_Code": cls_data.get("Classification_Code", "N/A"),
            "Flags": flags_str,
            "Info_Tags": ", ".join(result.info_tags),
            "Fields_To_Check": ", ".join(fields_to_check),
            "Decision_Required": decision_required,
            "Extracted_Values": extracted_values,
            "Auto_Corrections": result.auto_corrections,
            "Reviewer_Decision": "",
        }
        review_cards.append(card)

    # Save JSON
    json_path = output_dir / "human_review_queue.json"
    exc_str = ""
    try:
        json_path.write_text(
            json.dumps(review_cards, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
    except Exception as exc:
        exc_str = str(exc)
        logger.error("Failed to save review queue JSON: %s", exc_str)

    # Save Excel
    _save_review_excel(review_cards, output_dir / "human_review_queue.xlsx")

    return review_cards


def _save_review_excel(review_cards: list[dict], xlsx_path: Path) -> None:
    """Write human_review_queue.xlsx with the required columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Review Queue"

    headers = [
        "Paper_ID", "Title", "Flags",
        "Fields_To_Check", "Decision_Required", "Reviewer_Decision",
    ]

    # Bold header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # Data rows
    for row_idx, card in enumerate(review_cards, 2):
        ws.cell(row=row_idx, column=1, value=card.get("Paper_ID", ""))
        ws.cell(row=row_idx, column=2, value=card.get("Title", ""))
        ws.cell(row=row_idx, column=3, value=card.get("Flags", ""))
        ws.cell(row=row_idx, column=4, value=card.get("Fields_To_Check", ""))
        ws.cell(row=row_idx, column=5, value=card.get("Decision_Required", ""))
        ws.cell(row=row_idx, column=6, value="")  # Reviewer_Decision — left blank

    # Column widths
    col_widths = [15, 40, 50, 50, 80, 30]
    for col_idx, width in enumerate(col_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    exc_str: str = ""
    try:
        wb.save(xlsx_path)
    except Exception as exc:
        exc_str = str(exc)
        logger.error("Failed to save review queue Excel: %s", exc_str)


# ---------------------------------------------------------------------------
# Apply human decisions
# ---------------------------------------------------------------------------


def apply_human_decisions(review_path: str, extractions_dir: str) -> None:
    """Read a filled review Excel and write human decisions back to extraction JSONs.

    For each row where Reviewer_Decision is non-empty:
    - Loads {paper_id}_extraction.json
    - Adds Human_Review_Resolved=True and Human_Decision=<decision>
    - Saves back to disk
    """
    review_path_obj = Path(review_path)
    extractions_dir_obj = Path(extractions_dir)

    exc_str: str = ""
    try:
        wb = openpyxl.load_workbook(review_path_obj)
    except Exception as exc:
        exc_str = str(exc)
        logger.error("Failed to load review workbook: %s", exc_str)
        return

    ws = wb.active

    # Locate columns by header name
    header_row = [c.value for c in ws[1]]
    try:
        paper_id_col = header_row.index("Paper_ID") + 1
        decision_col = header_row.index("Reviewer_Decision") + 1
    except ValueError:
        logger.error(
            "Review Excel is missing required columns (Paper_ID or Reviewer_Decision)."
        )
        return

    applied = 0
    for row in ws.iter_rows(min_row=2):
        paper_id_cell = row[paper_id_col - 1]
        decision_cell = row[decision_col - 1]

        paper_id_val = paper_id_cell.value
        decision_val = decision_cell.value

        if not paper_id_val or not decision_val or not str(decision_val).strip():
            continue

        paper_id = str(paper_id_val).strip()
        decision = str(decision_val).strip()

        ext_path = extractions_dir_obj / f"{paper_id}_extraction.json"
        if not ext_path.exists():
            logger.warning("Extraction file not found for %s — skipping", paper_id)
            continue

        exc_str = ""
        try:
            extraction_data = json.loads(ext_path.read_text(encoding="utf-8"))
            extraction_data["Human_Review_Resolved"] = True
            extraction_data["Human_Decision"] = decision
            ext_path.write_text(
                json.dumps(extraction_data, indent=2, ensure_ascii=False),
                encoding="utf-8",
            )
            applied += 1
        except Exception as exc:
            exc_str = str(exc)
            logger.error("Failed to update extraction for %s: %s", paper_id, exc_str)

    print(f"Applied {applied} human decision(s).")


# ---------------------------------------------------------------------------
# Summary table
# ---------------------------------------------------------------------------


def _print_summary_table(results: list[QCResult]) -> None:
    total = len(results)
    passed = sum(1 for r in results if r.qc_passed)
    flagged = sum(1 for r in results if r.qc_flags)
    auto_corrected = sum(1 for r in results if r.auto_corrections)
    needs_review = sum(1 for r in results if r.requires_human_review)

    print("=" * 59)
    print("QC SUMMARY")
    print("=" * 59)
    print(f"Total papers processed : {total}")
    print(f"Passed QC (clean)      : {passed}")
    print(f"Flagged                : {flagged}")
    print(f"Auto-corrected         : {auto_corrected}  (papers with >=1 auto-correction)")
    print(f"Requires human review  : {needs_review}")
    print("=" * 59)

    flag_counter: Counter[str] = Counter()
    for r in results:
        for flag in r.qc_flags:
            flag_counter[flag] += 1

    if flag_counter:
        print("Flag breakdown:")
        for flag, count in sorted(flag_counter.items()):
            print(f"  {flag:<42}: {count}")


# ---------------------------------------------------------------------------
# Standalone runner
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    parser = argparse.ArgumentParser(description="Run QC on extracted papers.")
    parser.add_argument(
        "--extractions-dir",
        default=None,
        help="Directory containing *_extraction.json files (default: config.EXTRACTIONS_DIR)",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Directory for output artifacts (default: config.OUTPUT_DIR)",
    )
    parser.add_argument(
        "--apply-decisions",
        default=None,
        metavar="XLSX_PATH",
        help="Path to a filled review Excel; apply human decisions to extraction JSONs",
    )
    args = parser.parse_args()

    extractions_dir_arg = (
        Path(args.extractions_dir) if args.extractions_dir else config.EXTRACTIONS_DIR
    )
    output_dir_arg = Path(args.output_dir) if args.output_dir else config.OUTPUT_DIR

    if args.apply_decisions:
        apply_human_decisions(args.apply_decisions, str(extractions_dir_arg))
    else:
        qc_results = run_qc_all(extractions_dir_arg, output_dir_arg)
        _print_summary_table(qc_results)
        review_cards = generate_review_queue(qc_results, extractions_dir_arg, output_dir_arg)
        print(
            f"\nReview queue: {len(review_cards)} papers -> "
            f"{output_dir_arg}/human_review_queue.xlsx"
        )
