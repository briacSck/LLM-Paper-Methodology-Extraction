"""
Agent 4 -- Excel Exporter.

Assembles all per-paper extraction results into a single pandas DataFrame and
writes it to data/output/ as an Excel workbook (3 sheets) with a companion
summary report in Markdown.
"""

from __future__ import annotations

import json
import logging
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

_AGENTS_DIR = Path(__file__).resolve().parent
_PROJECT_ROOT = _AGENTS_DIR.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

import config
from schemas.extraction_schema import EXCEL_FIELDS

import anthropic
import pandas as pd
import openpyxl
import openpyxl.utils
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

SHEET1_COLUMNS: tuple[str, ...] = EXCEL_FIELDS + (
    "Classification_Code",
    "Eligible_Extraction",
    "QC_Flags",
    "Human_Review_Resolved",
)

SHEET2_COLUMNS: tuple[str, ...] = (
    "Paper_ID",
    "Timestamp",
    "Classification_Code",
    "Confidence_Score",
    "Time_Spent_Min",
    "All_Flags",
    "DV_Page",
    "IV_Page",
    "Missing_Data_Page",
    "Sample_Page",
    "Non_Obvious_Decisions",
    "Open_Questions",
    "Human_Review_Required",
    "Human_Decision",
    "QC_Auto_Corrections",
)

# ---------------------------------------------------------------------------
# Label lookup dicts
# ---------------------------------------------------------------------------

_MISSING_HANDLING_LABELS = {
    "1": "Listwise deletion",
    "2": "Pairwise deletion",
    "3": "Mean/mode substitution",
    "4": "Regression imputation",
    "5": "Multiple imputation (MI)",
    "6": "FIML",
    "7": "Single imputation",
    "8": "Complete case analysis",
    "9": "NR (not reported)",
    "10": "Other",
}

_MODEL_TYPE_LABELS = {
    "1": "OLS/Linear",
    "2": "Logit/Probit",
    "3": "Tobit",
    "4": "Count model",
    "5": "Panel FE/RE",
    "6": "2SLS/IV",
    "7": "SEM",
    "8": "HLM/Multilevel",
    "9": "Survival",
    "10": "Factor analysis",
    "11": "Other",
}

_REPLICATION_LABELS = {"1": "Yes", "2": "No", "3": "Partial"}

# ---------------------------------------------------------------------------
# Fill styles
# ---------------------------------------------------------------------------

_FILL_HEADER  = PatternFill(fgColor="BDD7EE", fill_type="solid")
_FILL_QC_FLAG = PatternFill(fgColor="FFFF99", fill_type="solid")
_FILL_MI_FIML = PatternFill(fgColor="C6EFCE", fill_type="solid")
_FILL_NR      = PatternFill(fgColor="FFC7CE", fill_type="solid")
_FILL_OTHER   = PatternFill(fgColor="FFEB9C", fill_type="solid")
_FONT_BOLD    = Font(bold=True)


# ---------------------------------------------------------------------------
# load_all_extractions
# ---------------------------------------------------------------------------

def load_all_extractions(
    extractions_dir: Optional[Path] = None,
    output_dir: Optional[Path] = None,
) -> list[dict]:
    """Load and merge classification + extraction data for all papers.

    Iterates all *_classification.json files (one per paper) and merges in
    the corresponding *_extraction.json where it exists.  Human review
    decisions from human_review_queue.xlsx are overlaid last.

    Returns a list of merged dicts, one per paper.
    """
    extractions_dir = Path(extractions_dir or config.EXTRACTIONS_DIR)
    output_dir = Path(output_dir or config.OUTPUT_DIR)

    cls_files = sorted(extractions_dir.glob("*_classification.json"))
    if not cls_files:
        logger.warning("No classification JSON files found in %s", extractions_dir)
        return []

    results: list[dict] = []

    for cls_path in cls_files:
        paper_id = cls_path.stem.replace("_classification", "")

        # Load classification
        exc_str: str = ""
        cls_data: dict = {}
        try:
            cls_data = json.loads(cls_path.read_text(encoding="utf-8"))
        except Exception as exc:
            exc_str = str(exc)
            logger.error("Failed to load classification %s: %s", cls_path.name, exc_str)
            continue

        # Load extraction (may not exist for non-EQR/MM papers)
        ext_path = extractions_dir / f"{paper_id}_extraction.json"
        exc_str = ""
        ext_data: dict = {}
        if ext_path.exists():
            try:
                ext_data = json.loads(ext_path.read_text(encoding="utf-8"))
            except Exception as exc:
                exc_str = str(exc)
                logger.error("Failed to load extraction %s: %s", ext_path.name, exc_str)
                # Treat as missing extraction rather than crash

        # Merge: extraction fields first, then classification fields on top for
        # Classification_Code and Eligible_Extraction
        merged: dict = {**ext_data}
        merged["Classification_Code"] = cls_data.get("Classification_Code", "NR")
        merged["Eligible_Extraction"] = cls_data.get("Eligible_Extraction", "NR")
        merged["_paper_id"] = paper_id
        merged["_has_extraction"] = bool(ext_data)
        merged["_cls_path"] = str(cls_path)
        merged["_ext_path"] = str(ext_path) if ext_path.exists() else ""
        merged["_cls_mtime"] = cls_path.stat().st_mtime
        merged["_ext_mtime"] = ext_path.stat().st_mtime if ext_path.exists() else None

        # Preserve classification sub-fields for log sheet
        merged["_cls_data"] = cls_data
        merged["_ext_data"] = ext_data

        results.append(merged)

    # Apply human review decisions from Excel
    review_path = output_dir / "human_review_queue.xlsx"
    if review_path.exists():
        exc_str = ""
        decisions: dict[str, str] = {}
        try:
            wb = openpyxl.load_workbook(review_path, read_only=True, data_only=True)
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            pid_col = headers.index("Paper_ID") if "Paper_ID" in headers else -1
            dec_col = headers.index("Reviewer_Decision") if "Reviewer_Decision" in headers else -1
            if pid_col >= 0 and dec_col >= 0:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    pid_val = row[pid_col]
                    dec_val = row[dec_col]
                    if pid_val and dec_val:
                        decisions[str(pid_val).strip()] = str(dec_val).strip()
        except Exception as exc:
            exc_str = str(exc)
            logger.error("Failed to load human review queue: %s", exc_str)

        if decisions:
            for merged in results:
                pid = merged["_paper_id"]
                if pid in decisions:
                    if not merged.get("Human_Review_Resolved"):
                        merged["Human_Review_Resolved"] = True
                        merged["Human_Decision"] = decisions[pid]

    logger.info("Loaded %d papers from %s", len(results), extractions_dir)
    return results


# ---------------------------------------------------------------------------
# build_extraction_dataframe
# ---------------------------------------------------------------------------

def build_extraction_dataframe(extractions: list[dict]) -> pd.DataFrame:
    """Build Sheet 1 DataFrame: one row per paper, all EXCEL_FIELDS + 4 pipeline cols."""
    rows: list[dict] = []

    for merged in extractions:
        has_ext = merged.get("_has_extraction", False)
        row: dict = {}

        if has_ext:
            # EQR/MM paper with extraction JSON
            for field in EXCEL_FIELDS:
                row[field] = str(merged.get(field, "NR"))

            qc = merged.get("qc_results", {})
            qc_flags_list = qc.get("qc_flags", []) if isinstance(qc, dict) else []
            row["QC_Flags"] = "; ".join(qc_flags_list) if qc_flags_list else ""

            row["Classification_Code"] = str(merged.get("Classification_Code", "NR"))
            row["Eligible_Extraction"] = str(merged.get("Eligible_Extraction", "NR"))
            row["Human_Review_Resolved"] = "Yes" if merged.get("Human_Review_Resolved") else "No"

        else:
            # Non-EQR paper — no extraction
            row["Paper_ID"]  = str(merged.get("Paper_ID", merged["_paper_id"]))
            row["Authors"]   = str(merged.get("Authors", "NR"))
            row["Year"]      = str(merged.get("Year", "NR"))
            row["Journal"]   = str(merged.get("Journal", "NR"))
            row["Title"]     = str(merged.get("Title", "NR"))

            # All remaining 60 EXCEL_FIELDS → "NA"
            for field in EXCEL_FIELDS:
                if field not in row:
                    row[field] = "NA"

            row["Classification_Code"]  = str(merged.get("Classification_Code", "NR"))
            row["Eligible_Extraction"]  = str(merged.get("Eligible_Extraction", "NR"))
            row["QC_Flags"]             = "NA"
            row["Human_Review_Resolved"] = "NA"

        rows.append(row)

    df = pd.DataFrame(rows, columns=list(SHEET1_COLUMNS))
    assert list(df.columns) == list(SHEET1_COLUMNS), (
        f"Column order mismatch: {list(df.columns)}"
    )
    return df


# ---------------------------------------------------------------------------
# build_log_dataframe
# ---------------------------------------------------------------------------

def build_log_dataframe(extractions: list[dict]) -> pd.DataFrame:
    """Build Sheet 2 DataFrame: one row per paper, pipeline audit log."""
    rows: list[dict] = []

    for merged in extractions:
        has_ext = merged.get("_has_extraction", False)
        cls_data = merged.get("_cls_data", {})
        ext_data = merged.get("_ext_data", {})
        qc = merged.get("qc_results", {})
        if not isinstance(qc, dict):
            qc = {}

        paper_id = merged.get("Paper_ID") or merged["_paper_id"]

        # Timestamp: prefer extraction file mtime, else classification mtime
        ext_mtime = merged.get("_ext_mtime")
        cls_mtime = merged.get("_cls_mtime", 0.0)
        if ext_mtime:
            timestamp = datetime.fromtimestamp(float(ext_mtime)).isoformat()
        else:
            timestamp = datetime.fromtimestamp(float(cls_mtime)).isoformat()

        # All_Flags: combine classification flag + extraction flag + QC flags
        qc_flags_list = qc.get("qc_flags", []) if qc else []
        flag_parts = [
            cls_data.get("Flag", ""),
            ext_data.get("Extraction_Flags", ""),
            "; ".join(qc_flags_list) if qc_flags_list else "",
        ]
        all_flags = "; ".join(f for f in flag_parts if f)

        # Time spent: from extraction data
        exc_str: str = ""
        try:
            time_min = round(float(ext_data.get("Time_Spent_Seconds", 0)) / 60, 2)
        except Exception as exc:
            exc_str = str(exc)
            time_min = 0.0

        row = {
            "Paper_ID": str(paper_id),
            "Timestamp": timestamp,
            "Classification_Code": str(merged.get("Classification_Code", "NR")),
            "Confidence_Score": str(cls_data.get("Confidence_Score", "NR")),
            "Time_Spent_Min": str(time_min),
            "All_Flags": all_flags,
            "DV_Page": str(
                merged.get("dv_page") or merged.get("Cat3_DV_Page", "NR")
            ),
            "IV_Page": str(
                merged.get("iv_page") or merged.get("Cat4_IV_Page", "NR")
            ),
            "Missing_Data_Page": str(
                merged.get("missing_data_page") or merged.get("Cat10_Missing_Page", "NR")
            ),
            "Sample_Page": str(
                merged.get("sample_page") or merged.get("Cat8_Sample_Page", "NR")
            ),
            "Non_Obvious_Decisions": str(merged.get("Extraction_Notes", "")),
            "Open_Questions": str(ext_data.get("Extraction_Flags", "")),
            "Human_Review_Required": str(qc.get("requires_human_review", False)),
            "Human_Decision": str(merged.get("Human_Decision", "")),
            "QC_Auto_Corrections": json.dumps(qc.get("auto_corrections", {})),
        }
        rows.append(row)

    df = pd.DataFrame(rows, columns=list(SHEET2_COLUMNS))
    return df


# ---------------------------------------------------------------------------
# build_summary_statistics
# ---------------------------------------------------------------------------

def build_summary_statistics(df: pd.DataFrame) -> dict:
    """Return {'tables': {name: DataFrame}, 'stats_dict': dict}."""
    import numpy as np  # noqa: F401 — optional, pandas uses it internally

    n_total = len(df)
    eqr_mask = df["Classification_Code"] == "EQR"
    n_eqr = int(eqr_mask.sum())
    eqr_df = df[eqr_mask].copy() if n_eqr > 0 else pd.DataFrame(columns=df.columns)

    # --- Table 1: Sample Composition ---
    t1 = df["Classification_Code"].value_counts().reset_index()
    t1.columns = ["Classification_Code", "Count"]
    t1["% of Total"] = (
        (t1["Count"] / n_total * 100).round(1).astype(str) + "%"
        if n_total > 0 else "0.0%"
    )

    # --- Table 2: Missing Data Handling (EQR only) ---
    if n_eqr > 0 and "Missing_Handling" in eqr_df.columns:
        mh_counts = eqr_df["Missing_Handling"].value_counts()
        t2 = mh_counts.reset_index()
        t2.columns = ["Code", "Count"]
        t2["Method"] = t2["Code"].map(_MISSING_HANDLING_LABELS).fillna("Unknown")
        t2["% of EQR"] = (t2["Count"] / n_eqr * 100).round(1).astype(str) + "%"
        t2 = t2[["Method", "Code", "Count", "% of EQR"]]
    else:
        t2 = pd.DataFrame(columns=["Method", "Code", "Count", "% of EQR"])

    # --- Table 3: Model Type (EQR only) ---
    if n_eqr > 0 and "Model_Type" in eqr_df.columns:
        mt_counts = eqr_df["Model_Type"].value_counts()
        t3 = mt_counts.reset_index()
        t3.columns = ["Code", "Count"]
        t3["Model"] = t3["Code"].map(_MODEL_TYPE_LABELS).fillna("Unknown")
        t3["% of EQR"] = (t3["Count"] / n_eqr * 100).round(1).astype(str) + "%"
        t3 = t3[["Model", "Code", "Count", "% of EQR"]]
    else:
        t3 = pd.DataFrame(columns=["Model", "Code", "Count", "% of EQR"])

    # --- Table 4: Reporting Transparency (EQR only) ---
    transparency_indicators = [
        "Missing_Mentioned", "Missing_Rate_Reported", "Missing_Justified",
        "Missing_Pattern_Tested", "Missing_Sensitivity", "Code_Available",
        "Software_Reported",
    ]
    t4_rows = []
    for ind in transparency_indicators:
        if n_eqr > 0 and ind in eqr_df.columns:
            yes_n = int((eqr_df[ind] == "1").sum())
            no_n  = int((eqr_df[ind] == "0").sum())
            yes_pct = f"{yes_n / n_eqr * 100:.1f}%"
            no_pct  = f"{no_n  / n_eqr * 100:.1f}%"
        else:
            yes_n = no_n = 0
            yes_pct = no_pct = "0.0%"
        t4_rows.append({
            "Indicator": ind,
            "Yes (N)": yes_n,
            "Yes (%)": yes_pct,
            "No (N)": no_n,
            "No (%)": no_pct,
        })
    t4 = pd.DataFrame(t4_rows)

    # --- Table 5: Replication Feasibility (EQR only) ---
    if n_eqr > 0 and "Replication_Feasibility" in eqr_df.columns:
        rep_counts = eqr_df["Replication_Feasibility"].value_counts()
        t5 = rep_counts.reset_index()
        t5.columns = ["Code", "Count"]
        t5["Label"] = t5["Code"].map(_REPLICATION_LABELS).fillna("Unknown")
        t5["% of EQR"] = (t5["Count"] / n_eqr * 100).round(1).astype(str) + "%"
        t5 = t5[["Label", "Code", "Count", "% of EQR"]]
    else:
        t5 = pd.DataFrame(columns=["Label", "Code", "Count", "% of EQR"])

    # --- stats_dict for LLM prompt ---
    classification_breakdown = (
        df["Classification_Code"].value_counts().to_dict()
        if n_total > 0 else {}
    )

    top_missing_method = "NR"
    if n_eqr > 0 and not t2.empty:
        top_row = t2.iloc[0]
        code = str(top_row["Code"])
        top_missing_method = _MISSING_HANDLING_LABELS.get(code, code)

    pct_mi_fiml = 0.0
    if n_eqr > 0 and "Missing_Handling" in eqr_df.columns:
        mi_fiml_n = int(((eqr_df["Missing_Handling"] == "5") | (eqr_df["Missing_Handling"] == "6")).sum())
        pct_mi_fiml = round(mi_fiml_n / n_eqr * 100, 1)

    pct_missing_mentioned = 0.0
    if n_eqr > 0 and "Missing_Mentioned" in eqr_df.columns:
        mentioned_n = int((eqr_df["Missing_Mentioned"] == "1").sum())
        pct_missing_mentioned = round(mentioned_n / n_eqr * 100, 1)

    n_flagged = 0
    if "QC_Flags" in df.columns:
        n_flagged = int((df["QC_Flags"].notna() & (df["QC_Flags"] != "") & (df["QC_Flags"] != "NA")).sum())

    n_reviewed = 0
    if "Human_Review_Resolved" in df.columns:
        n_reviewed = int((df["Human_Review_Resolved"] == "Yes").sum())

    stats_dict = {
        "n_total": n_total,
        "n_eqr": n_eqr,
        "classification_breakdown": classification_breakdown,
        "top_missing_method": top_missing_method,
        "pct_mi_fiml": pct_mi_fiml,
        "pct_missing_mentioned": pct_missing_mentioned,
        "n_flagged": n_flagged,
        "n_reviewed": n_reviewed,
        "missing_handling_table": t2.to_dict(orient="records") if not t2.empty else [],
        "model_type_table": t3.to_dict(orient="records") if not t3.empty else [],
        "transparency_table": t4.to_dict(orient="records") if not t4.empty else [],
        "replication_table": t5.to_dict(orient="records") if not t5.empty else [],
    }

    return {
        "tables": {
            "Sample_Composition": t1,
            "Missing_Data_Handling": t2,
            "Model_Type": t3,
            "Reporting_Transparency": t4,
            "Replication_Feasibility": t5,
        },
        "stats_dict": stats_dict,
    }


# ---------------------------------------------------------------------------
# generate_summary_report
# ---------------------------------------------------------------------------

def generate_summary_report(stats_dict: dict, client: anthropic.Anthropic) -> str:
    """Call the Anthropic API to generate a 600-800-word summary report."""
    # Build a text representation of the stats
    lines: list[str] = [
        "=== Extraction Summary Statistics ===",
        f"Total papers reviewed: {stats_dict['n_total']}",
        f"EQR (eligible for extraction): {stats_dict['n_eqr']}",
        f"Classification breakdown: {stats_dict['classification_breakdown']}",
        "",
        "--- Missing Data Handling (EQR papers) ---",
    ]

    for row in stats_dict.get("missing_handling_table", []):
        lines.append(f"  {row.get('Method','?')} (code {row.get('Code','?')}): "
                     f"N={row.get('Count','?')}, {row.get('% of EQR','?')}")

    lines += [
        "",
        f"Top method: {stats_dict['top_missing_method']}",
        f"MI or FIML used: {stats_dict['pct_mi_fiml']}%",
        f"Missing data mentioned at all: {stats_dict['pct_missing_mentioned']}%",
        "",
        "--- Model Type (EQR papers) ---",
    ]
    for row in stats_dict.get("model_type_table", []):
        lines.append(f"  {row.get('Model','?')} (code {row.get('Code','?')}): "
                     f"N={row.get('Count','?')}, {row.get('% of EQR','?')}")

    lines += [
        "",
        "--- Reporting Transparency (EQR papers) ---",
    ]
    for row in stats_dict.get("transparency_table", []):
        lines.append(f"  {row.get('Indicator','?')}: "
                     f"Yes={row.get('Yes (N)','?')} ({row.get('Yes (%)','?')}), "
                     f"No={row.get('No (N)','?')} ({row.get('No (%)','?')})")

    lines += [
        "",
        "--- Replication Feasibility (EQR papers) ---",
    ]
    for row in stats_dict.get("replication_table", []):
        lines.append(f"  {row.get('Label','?')} (code {row.get('Code','?')}): "
                     f"N={row.get('Count','?')}, {row.get('% of EQR','?')}")

    lines += [
        "",
        f"Papers flagged by QC: {stats_dict['n_flagged']}",
        f"Papers with human review resolved: {stats_dict['n_reviewed']}",
    ]

    stats_text = "\n".join(lines)

    prompt = (
        "Below are summary statistics from a structured extraction of "
        "management journal articles focusing on missing-data handling.\n\n"
        f"{stats_text}\n\n"
        "Write a 600-800-word academic summary with exactly four numbered sections:\n"
        "1. Sample Composition\n"
        "2. Missing Data Practices\n"
        "3. Methodological Patterns\n"
        "4. Recommendations for Full-Scale Extraction\n\n"
        "Cite all percentages and counts. Do not editorialize. "
        "Use precise, academic language."
    )

    exc_str: str = ""
    try:
        response = client.messages.create(
            model=config.CLAUDE_MODEL,
            max_tokens=1500,
            system=(
                "You are writing a methods section for an academic paper. "
                "Use precise language. Cite all percentages. Do not editorialize."
            ),
            messages=[{"role": "user", "content": prompt}],
        )
        return response.content[0].text
    except Exception as exc:
        exc_str = str(exc)
        logger.error("Failed to generate summary report: %s", exc_str)
        return ""


# ---------------------------------------------------------------------------
# export_to_excel
# ---------------------------------------------------------------------------

def export_to_excel(
    df_extraction: pd.DataFrame,
    df_log: pd.DataFrame,
    summary_stats: dict,
    output_path: Path,
) -> None:
    """Write a 3-sheet Excel workbook to output_path."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1 -- Extraction_Data                                          #
    # ------------------------------------------------------------------ #
    ws1 = wb.active
    ws1.title = "Extraction_Data"

    # Identify special column indices (1-based for openpyxl)
    col_names = list(SHEET1_COLUMNS)
    qc_flags_idx   = col_names.index("QC_Flags") + 1
    mh_idx         = col_names.index("Missing_Handling") + 1

    # Header row
    for col_num, col_name in enumerate(col_names, start=1):
        cell = ws1.cell(row=1, column=col_num, value=col_name)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_BOLD

    # Data rows
    for row_num, (_, df_row) in enumerate(df_extraction.iterrows(), start=2):
        for col_num, col_name in enumerate(col_names, start=1):
            val = df_row[col_name]
            cell = ws1.cell(row=row_num, column=col_num, value=str(val))

            # Conditional fill: QC_Flags
            if col_num == qc_flags_idx and val and val not in ("", "NA"):
                cell.fill = _FILL_QC_FLAG

            # Conditional fill: Missing_Handling
            if col_num == mh_idx:
                if val in ("5", "6"):
                    cell.fill = _FILL_MI_FIML
                elif val == "9":
                    cell.fill = _FILL_NR
                elif val == "10":
                    cell.fill = _FILL_OTHER

    # Freeze panes: first unfrozen cell is B2 (freezes row 1 + column A)
    ws1.freeze_panes = "B2"

    # Auto-fit column widths
    for col_idx, col_name in enumerate(col_names, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        col_values = [col_name] + list(df_extraction[col_name].astype(str))
        max_len = min(max((len(v) for v in col_values), default=0), 50) + 2
        ws1.column_dimensions[col_letter].width = max_len

    # ------------------------------------------------------------------ #
    # Sheet 2 -- Extraction_Log                                           #
    # ------------------------------------------------------------------ #
    ws2 = wb.create_sheet("Extraction_Log")
    log_col_names = list(SHEET2_COLUMNS)

    # Header
    for col_num, col_name in enumerate(log_col_names, start=1):
        cell = ws2.cell(row=1, column=col_num, value=col_name)
        cell.font = _FONT_BOLD

    # Data rows
    for row_num, (_, df_row) in enumerate(df_log.iterrows(), start=2):
        for col_num, col_name in enumerate(log_col_names, start=1):
            val = df_row[col_name]
            ws2.cell(row=row_num, column=col_num, value=str(val))

    # Auto-fit
    for col_idx, col_name in enumerate(log_col_names, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        col_values = [col_name] + list(df_log[col_name].astype(str))
        max_len = min(max((len(v) for v in col_values), default=0), 50) + 2
        ws2.column_dimensions[col_letter].width = max_len

    # ------------------------------------------------------------------ #
    # Sheet 3 -- Summary_Statistics                                       #
    # ------------------------------------------------------------------ #
    ws3 = wb.create_sheet("Summary_Statistics")
    tables = summary_stats.get("tables", {})
    table_titles = {
        "Sample_Composition":      "Table 1 -- Sample Composition",
        "Missing_Data_Handling":   "Table 2 -- Missing Data Handling (EQR only)",
        "Model_Type":              "Table 3 -- Model Type (EQR only)",
        "Reporting_Transparency":  "Table 4 -- Reporting Transparency (EQR only)",
        "Replication_Feasibility": "Table 5 -- Replication Feasibility (EQR only)",
    }

    current_row = 1
    for table_name, title in table_titles.items():
        tdf = tables.get(table_name, pd.DataFrame())

        # Table title
        title_cell = ws3.cell(row=current_row, column=1, value=title)
        title_cell.font = _FONT_BOLD
        current_row += 1

        if tdf.empty:
            ws3.cell(row=current_row, column=1, value="(no data)")
            current_row += 2
            continue

        # Header row
        for col_num, col_name in enumerate(tdf.columns, start=1):
            cell = ws3.cell(row=current_row, column=col_num, value=col_name)
            cell.font = _FONT_BOLD

        # Track column widths
        col_widths: dict[int, int] = {}
        for col_num, col_name in enumerate(tdf.columns, start=1):
            col_widths[col_num] = len(str(col_name))

        current_row += 1

        # Data rows
        for _, trow in tdf.iterrows():
            for col_num, col_name in enumerate(tdf.columns, start=1):
                val = str(trow[col_name])
                ws3.cell(row=current_row, column=col_num, value=val)
                col_widths[col_num] = max(col_widths.get(col_num, 0), len(val))
            current_row += 1

        # Auto-fit column widths for this table
        for col_num, width in col_widths.items():
            col_letter = openpyxl.utils.get_column_letter(col_num)
            ws3.column_dimensions[col_letter].width = min(width + 2, 50)

        # Blank row between tables
        current_row += 1

    wb.save(str(output_path))
    logger.info("Excel workbook saved to %s", output_path)


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main(
    extractions_dir: Optional[Path] = None,
    output_dir: Optional[Path] = None,
    api_key: Optional[str] = None,
) -> None:
    """Run Phase 5: assemble extraction data and export to Excel + Markdown."""
    extractions_dir = Path(extractions_dir or config.EXTRACTIONS_DIR)
    output_dir = Path(output_dir or config.OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)

    api_key = api_key or config.ANTHROPIC_API_KEY

    extractions = load_all_extractions(extractions_dir, output_dir)
    if not extractions:
        logger.warning("No extractions found; Excel output will be empty.")

    df_extraction = build_extraction_dataframe(extractions)
    df_log        = build_log_dataframe(extractions)
    summary_stats = build_summary_statistics(df_extraction)

    stats_dict = summary_stats["stats_dict"]
    n_total   = stats_dict["n_total"]
    n_eqr     = stats_dict["n_eqr"]
    n_flagged = stats_dict["n_flagged"]
    n_reviewed = stats_dict["n_reviewed"]

    # Generate narrative report via API if key is available
    report_md: str = ""
    if api_key:
        try:
            client = anthropic.Anthropic(api_key=api_key)
            report_md = generate_summary_report(stats_dict, client)
        except Exception as exc:
            exc_str: str = str(exc)
            logger.error("Could not initialise Anthropic client: %s", exc_str)
    else:
        logger.warning("ANTHROPIC_API_KEY not set; skipping summary report generation.")

    output_xlsx = output_dir / "extraction_output.xlsx"
    export_to_excel(df_extraction, df_log, summary_stats, output_xlsx)

    report_path = output_dir / "summary_report.md"
    if report_md:
        report_path.write_text(report_md, encoding="utf-8")
        logger.info("Summary report saved to %s", report_path)
    else:
        report_path.write_text("# Summary Report\n\n(Not generated)\n", encoding="utf-8")

    print(
        f"Export complete: {n_total} papers, {n_eqr} EQR extracted, "
        f"{n_flagged} flagged, {n_reviewed} human-reviewed"
    )
    print(f"Output: {output_xlsx}")
    print(f"Report: {report_path}")


# ---------------------------------------------------------------------------
# Standalone runner
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )

    parser = argparse.ArgumentParser(description="Agent 4 -- Excel Exporter")
    parser.add_argument(
        "--extractions-dir",
        type=Path,
        default=None,
        help="Directory containing *_classification.json and *_extraction.json files.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help="Directory for Excel + Markdown output.",
    )
    args = parser.parse_args()

    main(
        extractions_dir=args.extractions_dir,
        output_dir=args.output_dir,
    )
